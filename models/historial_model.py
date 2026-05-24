import os
from datetime import datetime, timedelta

HISTORIAL_DIR  = 'historial'
DIAS_HISTORIAL = 60   # Retencion de 60 dias (2 meses)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_OK = True
except ImportError:
    EXCEL_OK = False

CABECERAS = [
    'Fecha y hora','Nombre','Sala','Ubicacion','Marca','Modelo',
    'Estado','Bateria %','V. Entrada','V. Salida',
    'Carga W','Carga %','Temperatura C','Autonomia min'
]

FILLS = {
    'En linea'     : 'E8F5E9',
    'En bateria'   : 'FFEBEE',
    'Falla'        : 'FFEBEE',
    'Bypass'       : 'FFEBEE',
    'Sin respuesta': 'F5F5F5',
}

def nombre_archivo_dia() -> str:
    """Genera el nombre del archivo Excel del dia actual."""
    hoy = datetime.now()
    return os.path.join(HISTORIAL_DIR, f'historial_{hoy.strftime("%Y-%m-%d")}.xlsx')

def guardar_historial(datos: list[dict]) -> None:
    if not EXCEL_OK:
        return
    os.makedirs(HISTORIAL_DIR, exist_ok=True)
    archivo = nombre_archivo_dia()
    ahora   = datetime.now().strftime('%d/%m/%Y %H:%M:%S')

    if os.path.exists(archivo):
        wb = openpyxl.load_workbook(archivo)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Historial'
        ws.append(CABECERAS)
        for col in range(1, len(CABECERAS) + 1):
            c = ws.cell(1, col)
            c.fill = PatternFill('solid', fgColor='1565C0')
            c.font = Font(color='FFFFFF', bold=True, size=10)
            c.alignment = Alignment(horizontal='center')
        ws.freeze_panes = 'A2'

    for u in datos:
        fila = [
            ahora,
            u.get('nombre',''),    u.get('sala',''),
            u.get('ubicacion',''), u.get('marca','').upper(),
            u.get('modelo',''),    u.get('estado',''),
            u.get('bateria_pct',''),    u.get('voltaje_entrada',''),
            u.get('voltaje_salida',''), u.get('carga_w',''),
            u.get('carga_pct',''),      u.get('temperatura',''),
            u.get('autonomia_min',''),
        ]
        ws.append(fila)
        color = FILLS.get(u.get('estado',''), FILLS['Sin respuesta'])
        for col in range(1, len(CABECERAS) + 1):
            ws.cell(ws.max_row, col).fill = PatternFill('solid', fgColor=color)

    anchos = [18,12,8,18,8,16,14,10,10,10,10,10,14,14]
    for i, a in enumerate(anchos, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = a

    wb.save(archivo)

def limpiar_historial_viejo() -> None:
    if not os.path.exists(HISTORIAL_DIR):
        return
    limite = (datetime.now() - timedelta(days=DIAS_HISTORIAL)).timestamp()
    for f in os.listdir(HISTORIAL_DIR):
        ruta = os.path.join(HISTORIAL_DIR, f)
        if os.path.getmtime(ruta) < limite:
            os.remove(ruta)

def listar_historiales() -> list[str]:
    if not os.path.exists(HISTORIAL_DIR):
        return []
    return sorted(
        [f for f in os.listdir(HISTORIAL_DIR) if f.endswith('.xlsx')],
        reverse=True
    )

def leer_historial(archivo: str) -> dict:
    if not EXCEL_OK or not archivo:
        return {'cabeceras': [], 'filas': []}
    ruta = os.path.join(HISTORIAL_DIR, archivo)
    if not os.path.exists(ruta):
        return {'cabeceras': [], 'filas': []}
    wb    = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
    ws    = wb.active
    filas = list(ws.iter_rows(values_only=True))
    if not filas:
        return {'cabeceras': [], 'filas': []}
    cab   = [str(c) if c else '' for c in filas[0]]
    datos = [[str(v) if v is not None else '' for v in f] for f in filas[1:]]
    return {'cabeceras': cab, 'filas': datos, 'total': len(datos)}